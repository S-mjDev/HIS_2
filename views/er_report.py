import os
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from models.patient_database import PatientDatabase
from utils.theme import T, FONT
from utils.widgets import mk_entry, page_header
from utils.validators import validate_date

try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ── Colour palette — ALL must be 8-char aRGB ─────────────
XL_ACCENT      = "FF2563EB"
XL_ACCENT_DARK = "FF1D4ED8"
XL_SIDEBAR     = "FF1A2332"
XL_HEADING_BG  = "FFF1F5F9"
XL_ROW_EVEN    = "FFF8FAFC"
XL_ROW_ODD     = "FFFFFFFF"
XL_BORDER      = "FFE2E8F0"
XL_TEXT        = "FF1E293B"
XL_MUTED       = "FF64748B"
XL_WHITE       = "FFFFFFFF"
XL_SUCCESS     = "FF059669"
XL_WARNING     = "FFD97706"
XL_DANGER      = "FFDC2626"


def _border(color=XL_BORDER, style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=XL_TEXT, size=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _arrival_time_color(time_str):
    """Return 8-char aRGB hex based on arrival time range.
    12:00 AM – 6:59 AM  → RED
    7:00 AM  – 2:59 PM  → BLACK
    3:00 PM  – 10:59 PM → BLUE
    11:00 PM – 11:59 PM → RED
    """
    if not time_str:
        return XL_TEXT
    try:    
        ts = str(time_str).strip().upper()
        if "AM" in ts or "PM" in ts:
            t = datetime.strptime(ts, "%I:%M %p")
        else:
            t = datetime.strptime(ts, "%H:%M")
        total_minutes = t.hour * 60 + t.minute
        if 0 <= total_minutes <= 419:        # 12:00 AM – 6:59 AM
            return "FFDC2626"                # RED
        elif 420 <= total_minutes <= 899:    # 7:00 AM  – 2:59 PM
            return XL_TEXT                   # BLACK
        elif 900 <= total_minutes <= 1379:   # 3:00 PM  – 10:59 PM
            return "FF2563EB"                # BLUE
        else:                                # 11:00 PM – 11:59 PM
            return "FFDC2626"                # RED
    except ValueError:
        return XL_TEXT


def _service_type_color(service_str):
    """Return (fill_aRGB, text_aRGB) based on Type of Service."""
    colors = {
        "MEDICINE":   ("FFD1FAE5", "FF065F46"),  # green
        "SURGICAL":   ("FFFEE2E2", "FFB91C1C"),  # red
        "OB-GYNE":    ("FFFDF4FF", "FF86198F"),  # magenta
        "PEDIATRICS": ("FFFEFCE8", "FF854D0E"),  # yellow
        "OTHERS":     ("FFF1F5F9", "FF475569"),  # grey
    }
    key = str(service_str).strip().upper() if service_str else ""
    return colors.get(key, ("FFFFFFFF", "FF1E293B"))


def build_er_report_page(parent, page_refreshers=None):
    frame = Frame(parent, bg=T["bg"])
    page_header(frame, "ER Patient Report",
                "Export ER patient registration data to Excel")

    db_container = {"instance": PatientDatabase()}

    content = Frame(frame, bg=T["bg"])
    content.pack(fill=BOTH, expand=True, padx=24, pady=16)

    def _normalize_date(date_text):
        for fmt in ["%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"]:
            try:
                return datetime.strptime(date_text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None

    def refresh_database():
        try:
            db_container["instance"] = PatientDatabase()
            return True
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to refresh database:\n{e}")
            return False

    # ── Filter card ───────────────────────────────────────
    filter_card = Frame(content, bg=T["panel"],
                        highlightthickness=1, highlightbackground=T["border"])
    filter_card.pack(fill=X, pady=(0, 18))
    fb = Frame(filter_card, bg=T["panel"])
    fb.pack(fill=X, padx=20, pady=16)

    Label(fb, text="EXPORT FILTER", font=FONT["tag"],
          bg=T["panel"], fg=T["accent"]).grid(row=0, column=0, columnspan=4, sticky=W)
    Label(fb, text="Exports ER-registered patients only.", font=FONT["body"],
          bg=T["panel"], fg=T["muted"]).grid(row=1, column=0, columnspan=4,
                                              sticky=W, pady=(4, 12))
    try:
        from tkcalendar import DateEntry as _DateEntry
        _HAS_CAL = True
    except ImportError:
        _HAS_CAL = False

    Label(fb, text="START DATE", font=FONT["tag"],
        bg=T["panel"], fg=T["muted"]).grid(row=2, column=0, sticky=W, pady=(0, 4))

    if _HAS_CAL:
        from_entry = _DateEntry(fb, width=16, date_pattern="mm-dd-yyyy",
                                font=FONT["body"],
                                background=T["accent"], foreground=T["white"],
                                headersbackground=T["accent"],
                                selectbackground=T["accent"],
                                normalforeground=T["text"],
                                weekendforeground=T["danger"],
                                borderwidth=1)
        from_entry.grid(row=3, column=0, sticky=W, ipady=4)
    else:
        from_entry = mk_entry(fb, width=18)
        from_entry.grid(row=3, column=0, sticky=W)

    Label(fb, text="END DATE", font=FONT["tag"],
        bg=T["panel"], fg=T["muted"]).grid(row=2, column=1, sticky=W,
                                            pady=(0, 4), padx=(24, 0))
    if _HAS_CAL:
        to_entry = _DateEntry(fb, width=16, date_pattern="mm-dd-yyyy",
                            font=FONT["body"],
                            background=T["accent"], foreground=T["white"],
                            headersbackground=T["accent"],
                            selectbackground=T["accent"],
                            normalforeground=T["text"],
                            weekendforeground=T["danger"],
                            borderwidth=1)
        to_entry.grid(row=3, column=1, sticky=W, padx=(24, 0), ipady=4)
    else:
        to_entry = mk_entry(fb, width=18)
        to_entry.grid(row=3, column=1, sticky=W, padx=(24, 0))

    Label(fb, text="Click the calendar icon to pick a date",
        font=FONT["small"], bg=T["panel"],
        fg=T["muted"]).grid(row=4, column=0, columnspan=3, sticky=W, pady=(6, 0))

    if openpyxl is None:
        Label(content, text="⚠  openpyxl is required.  Run: pip install openpyxl",
              font=FONT["body"], bg=T["bg"], fg=T["danger"]).pack(anchor=W, pady=12)
        return frame

    # ── Excel export ──────────────────────────────────────
    def export_to_excel():
        if not refresh_database():
            return

        start_raw  = from_entry.get().strip()
        end_raw    = to_entry.get().strip()
        start_date = end_date = None

        # For DateEntry, get_date() returns a datetime.date object directly
        if hasattr(from_entry, "get_date") and from_entry.get().strip():
            start_date = from_entry.get_date().strftime("%Y-%m-%d")
            start_raw  = start_date
        else:
            start_raw  = from_entry.get().strip()

        if hasattr(to_entry, "get_date") and to_entry.get().strip():
            end_date = to_entry.get_date().strftime("%Y-%m-%d")
            end_raw  = end_date
        else:
            end_raw = to_entry.get().strip()

        patients = db_container["instance"].get_patients(
            er_only=True, start_date=start_date, end_date=end_date)
        if not patients:
            messagebox.showinfo("No Data",
                "No ER patient records found for the selected range.")
            return
        
        admissions = db_container["instance"].get_all_admissions( start_date=start_date, end_date=end_date)

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            title="Save ER Patient Report"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ER Patient Report"

            # ── Title block ───────────────────────────────
            date_label = ""
            if start_date and end_date:
                date_label = f"  {start_date}  to  {end_date}"
            elif start_date:
                date_label = f"  From  {start_date}"
            elif end_date:
                date_label = f"  Up to  {end_date}"

            ws.merge_cells("A1:Z1")
            ws["A1"].value     = "HOSPITAL INFORMATION SYSTEM"
            ws["A1"].font      = _font(bold=True, color=XL_WHITE, size=14)
            ws["A1"].fill      = _fill(XL_SIDEBAR)
            ws["A1"].alignment = _align("center")

            ws.merge_cells("A2:Z2")
            ws["A2"].value     = f"ER PATIENT REPORT{date_label}"
            ws["A2"].font      = _font(bold=True, color=XL_WHITE, size=11)
            ws["A2"].fill      = _fill(XL_ACCENT)
            ws["A2"].alignment = _align("center")

            ws.merge_cells("A3:Z3")
            ws["A3"].value = (
                f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
                f"  |  Total Records: {len(patients)}"
            )
            ws["A3"].font      = _font(italic=True, color=XL_MUTED, size=9)
            ws["A3"].fill      = _fill(XL_HEADING_BG)
            ws["A3"].alignment = _align("center")

            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 22
            ws.row_dimensions[3].height = 16
            ws.append([])  # blank row 4

            # ── Column headers (row 5) ────────────────────
            headers = [
                ("Case No.",         10),
                ("Patient ID",       12),
                ("Arrival Time",     14),
                ("First Name",       16),
                ("Middle Name",      14),
                ("Last Name",        16),
                ("Age",               6),
                ("Gender",            9),
                ("Civil Status",     13),
                ("Birth Date",       14),
                ("Birth Place",      16),
                ("Nationality",      13),
                ("Barangay",         18),
                ("Municipality",     16),
                ("Province",         14),
                ("Diagnosis",        30),
                ("Phone Number",     16),
                ("Type of Service",  18),
                ("Referral To",      18),
                ("Seen by Doctor",   16),
                ("Disposition",      16),
                ("Time if Admit",    14),
                ("Doctor",           18),
                ("Medical History",  22),
                ("Registered By",    14),
                ("Registration Date",20),
            ]

            header_row = 5
            for col_idx, (label, width) in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=label)
                cell.font      = _font(bold=True, color=XL_WHITE, size=9)
                cell.fill      = _fill(XL_ACCENT)
                cell.alignment = _align("center")
                cell.border    = _border(XL_ACCENT_DARK)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[header_row].height = 20

            # ── Data rows ─────────────────────────────────
            for row_idx, patient in enumerate(patients.values(),
                                              start=header_row + 1):
                is_even  = (row_idx - header_row) % 2 == 0
                row_fill = _fill(XL_ROW_EVEN if is_even else XL_ROW_ODD)

                row_data = [
                    patient.get("case_number",      ""),
                    patient.get("patient_id",       ""),
                    patient.get("arrival_time",     ""),
                    patient.get("first_name",       ""),
                    patient.get("middle_name",      ""),
                    patient.get("last_name",        ""),
                    patient.get("age",              ""),
                    patient.get("gender",           ""),
                    patient.get("civil_status",     ""),
                    patient.get("birth_date",       ""),
                    patient.get("birth_place",      ""),
                    patient.get("nationality",      ""),
                    patient.get("barangay",         ""),
                    patient.get("municipality",     ""),
                    patient.get("province",         ""),
                    patient.get("diagnosis",        ""),
                    patient.get("phone",            ""),
                    patient.get("service_type",     ""),
                    patient.get("referred_to",      ""),
                    patient.get("seen_by_doctor",   ""),
                    patient.get("disposition",      ""),
                    patient.get("time_if_admit",    ""),
                    patient.get("doctor",           ""),
                    patient.get("medical_history",  ""),
                    patient.get("registered_by",    ""),
                    patient.get("registration_date",""),
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill      = row_fill
                    cell.border    = _border()
                    cell.font      = _font(size=9)
                    # Names left-aligned, everything else centered
                    cell.alignment = _align(
                        "left" if col_idx in (4, 5, 6, 16, 24) else "center",
                        wrap=col_idx in (16, 24)   # Diagnosis & Medical History
                    )

                    # Case Number — blue bold
                    if col_idx == 1 and value:
                        cell.font = _font(bold=True, color=XL_ACCENT, size=9)

                    # Arrival Time — colour by time range
                    elif col_idx == 3 and value:
                        color = _arrival_time_color(str(value))
                        cell.font = _font(bold=True, color=color, size=9)

                    # Type of Service — colour by service type
                    elif col_idx == 18 and value:
                        fill_hex, text_hex = _service_type_color(str(value))
                        cell.fill      = _fill(fill_hex)
                        cell.font      = _font(bold=True, color=text_hex, size=9)
                        cell.alignment = _align("center")
                    
                    elif col_idx in (8, 9, 19) and value:
                        cell.font = _font(bold=True, color=XL_DANGER, size=9)

                # Dynamic row height for wrapped columns
                max_lines = 1
                for ci, val in enumerate(row_data, start=1):
                    if val and ci in (16, 24):
                        max_lines = max(max_lines, len(str(val)) // 45 + 1)
                ws.row_dimensions[row_idx].height = max(16, max_lines * 15)

            # ── Auto-fit column widths ─────────────────────
            for col_idx, (label, min_width) in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_length = len(label)
                for r in range(header_row + 1, header_row + len(patients) + 1):
                    val = ws.cell(row=r, column=col_idx).value
                    if val:
                        lines      = str(val).split("\n")
                        max_length = max(max_length, max(len(l) for l in lines))
                fitted = min(max_length + 4, 60)
                ws.column_dimensions[col_letter].width = max(fitted, min_width)

            # ── Freeze panes & auto-filter ─────────────────
            ws.freeze_panes   = "A6"
            ws.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
            )

            # ── Summary footer ─────────────────────────────
            footer_row = header_row + len(patients) + 2
            ws.merge_cells(
                f"A{footer_row}:{get_column_letter(len(headers))}{footer_row}"
            )
            footer            = ws[f"A{footer_row}"]
            footer.value      = (
                f"END OF REPORT  |  Total: {len(patients)} record(s)  |  "
                f"Exported: {datetime.now().strftime('%B %d, %Y')}"
            )
            footer.font       = _font(italic=True, color=XL_MUTED, size=8)
            footer.fill       = _fill(XL_HEADING_BG)
            footer.alignment  = _align("center")
            footer.border     = _border()

            # ── Print settings ─────────────────────────────
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage   = True
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 0

                        # ── Sheet protection (lock for editing) ────────
            ws.protection.sheet         = True
            ws.protection.password      = "qphn2025"
            ws.protection.selectLockedCells   = False  # allow selecting locked cells
            ws.protection.selectUnlockedCells = False  # allow selecting unlocked cells
            ws.protection.formatCells         = True   # block formatting
            ws.protection.formatColumns       = True   # block column resize
            ws.protection.formatRows          = True   # block row resize
            ws.protection.insertRows          = True   # block insert rows
            ws.protection.insertColumns       = True   # block insert columns
            ws.protection.deleteRows          = True   # block delete rows
            ws.protection.deleteColumns       = True   # block delete columns
            ws.protection.sort                = True   # block sorting
            ws.protection.autoFilter          = False   # block filter changes

            # ══════════════════════════════════════════════
            # SHEET 2 — DISCHARGE TURNAROUND TIME (TAT)
            # ══════════════════════════════════════════════
            ws2 = wb.create_sheet(title="Discharge TAT")

            for row_idx, admission in enumerate(admissions.values(),
                                              start=header_row + 1):
                is_even  = (row_idx - header_row) % 2 == 0
                row_fill = _fill(XL_ROW_EVEN if is_even else XL_ROW_ODD)


            # Only include patients with disposition = DISCHARGED
            # and who have time_if_admit (doctor's discharge order time)
            discharged = [
                admission.get("admission_id") for admission in admissions.values()
            ]

            # ── Title block ───────────────────────────────
            TAT_COLS = 6
            TAT_LAST = get_column_letter(TAT_COLS)

            ws2.merge_cells(f"A1:{TAT_LAST}1")
            ws2["A1"].value     = "HOSPITAL INFORMATION SYSTEM"
            ws2["A1"].font      = _font(bold=True, color=XL_WHITE, size=14)
            ws2["A1"].fill      = _fill(XL_SIDEBAR)
            ws2["A1"].alignment = _align("center")

            ws2.merge_cells(f"A2:{TAT_LAST}2")
            ws2["A2"].value     = f"DISCHARGE TURNAROUND TIME (TAT) REPORT{date_label}"
            ws2["A2"].font      = _font(bold=True, color=XL_WHITE, size=11)
            ws2["A2"].fill      = _fill("FFEA580C")   # orange header
            ws2["A2"].alignment = _align("center")

            ws2.merge_cells(f"A3:{TAT_LAST}3")
            ws2["A3"].value = (
                f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
                f"  |  Total Discharged: {len(discharged)}"
                f"  |  TAT = Time of Actual Discharge − Time of Dr's Order"
            )
            ws2["A3"].font      = _font(italic=True, color=XL_MUTED, size=9)
            ws2["A3"].fill      = _fill(XL_HEADING_BG)
            ws2["A3"].alignment = _align("center")

            ws2.row_dimensions[1].height = 26
            ws2.row_dimensions[2].height = 22
            ws2.row_dimensions[3].height = 16
            ws2.append([])   # blank row 4

            # ── Column headers (row 5) ────────────────────
            tat_headers = [
                ("Date of Discharge",               18),
                ("Patient Name",                    28),
                ("Time of Discharge (Dr's Orders)", 28),
                ("Time of Actual Discharge",        28),
                ("Remarks",                         30),
                ("TAT",                             14),
            ]

            tat_header_row = 5
            for col_idx, (label, width) in enumerate(tat_headers, start=1):
                cell           = ws2.cell(row=tat_header_row, column=col_idx, value=label)
                cell.font      = _font(bold=True, color=XL_WHITE, size=9)
                cell.fill      = _fill("FFEA580C")   # orange
                cell.alignment = _align("center", wrap=True)
                cell.border    = _border("FFD4520A")
                ws2.column_dimensions[get_column_letter(col_idx)].width = width

            ws2.row_dimensions[tat_header_row].height = 28

            def _parse_time(time_str):
                """Parse time string, return datetime or None."""
                if not time_str:
                    return None
                ts = str(time_str).strip().upper()
                for fmt in ["%I:%M %p", "%H:%M:%S", "%H:%M"]:
                    try:
                        return datetime.strptime(ts, fmt)
                    except ValueError:
                        continue
                return None

            def _compute_tat(dr_order_time_str, actual_discharge_str):
                """
                Compute TAT = actual_discharge - dr_order_time.
                Returns formatted string like '+01:30:00' or '-00:20:00'.
                Negative TAT means patient left before the order was written (data issue).
                """
                t1 = _parse_time(dr_order_time_str)
                t2 = _parse_time(actual_discharge_str)
                if not t1 or not t2:
                    return "—"
                try:
                    # Use today's date as base since only time is stored
                    base = datetime.today().replace(
                        hour=0, minute=0, second=0, microsecond=0)
                    dt1 = base.replace(hour=t1.hour, minute=t1.minute, second=t1.second)
                    dt2 = base.replace(hour=t2.hour, minute=t2.minute, second=t2.second)
                    delta = dt2 - dt1
                    total_seconds = int(delta.total_seconds())
                    sign   = "-" if total_seconds < 0 else ""
                    total_seconds = abs(total_seconds)
                    hours  = total_seconds // 3600
                    minutes= (total_seconds % 3600) // 60
                    seconds= total_seconds % 60
                    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"
                except Exception:
                    return "—"

            def _tat_color(tat_str):
                """Color TAT cell: green ≤30min, orange ≤60min, red >60min, grey if N/A."""
                if not tat_str or tat_str == "—":
                    return XL_MUTED
                try:
                    negative = tat_str.startswith("-")
                    clean    = tat_str.lstrip("-")
                    parts    = clean.split(":")
                    total_min = int(parts[0]) * 60 + int(parts[1])
                    if negative:
                        return "FF7C3AED"   # purple — negative TAT (unusual)
                    elif total_min <= 30:
                        return XL_SUCCESS   # green — fast
                    elif total_min <= 60:
                        return XL_WARNING   # orange — acceptable
                    else:
                        return XL_DANGER    # red — slow
                except Exception:
                    return XL_MUTED

            # ── Data rows ─────────────────────────────────
            for row_idx, (pid, a) in enumerate(
                    admissions.items(), start=tat_header_row + 1):

                is_even  = (row_idx - tat_header_row) % 2 == 0
                row_fill = _fill(XL_ROW_EVEN if is_even else XL_ROW_ODD)

                full_name = " ".join(filter(None, [
                    a.get("first_name", ""),
                    a.get("middle_name", ""),
                    a.get("last_name", "")
                ])).title()

                # Date of discharge = registration_date date part
                reg_date = a.get("discharge_date", "")
                dis_date = str(reg_date)[:10] if reg_date else ""

                # Dr's order time = time_of_discharge_dr_order field
                dr_order_time   = a.get("time_of_discharged_dr_order", "")
                dr_order_time_str = str(dr_order_time)[11:19] if dr_order_time else ""
                dr_order_time   = dr_order_time_str if dr_order_time_str else ""
                # Actual discharge = arrival_time used as proxy
                # (update this if you add a separate actual_discharge_time field)
                dr_remarks = a.get("remarks", "") or ""
                actual_dis_time = a.get("discharge_date", "")
                dis_time_str = str(actual_dis_time)[11:19] if actual_dis_time else ""
                actual_dis_time = dis_time_str if dis_time_str else ""

                tat_str   = _compute_tat(dr_order_time, actual_dis_time)
                tat_color = _tat_color(tat_str)

                row_data = [
                    dis_date,
                    full_name,
                    dr_order_time,
                    actual_dis_time,
                    dr_remarks,
                    a.get("medical_history", "") or a.get("referred_to", "") or "",
                    tat_str,
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell           = ws2.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill      = row_fill
                    cell.border    = _border()
                    cell.font      = _font(size=9)
                    cell.alignment = _align(
                        "left" if col_idx in (2, 5) else "center",
                        wrap=col_idx == 5
                    )

                    # TAT column — colour by duration
                    if col_idx == 6:
                        cell.font = _font(bold=True, color=tat_color, size=9)

                ws2.row_dimensions[row_idx].height = 16

            # ── Auto-fit ──────────────────────────────────
            for col_idx, (label, min_w) in enumerate(tat_headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len    = len(label)
                for r in range(tat_header_row + 1,
                               tat_header_row + len(admissions) + 1):
                    val = ws2.cell(row=r, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws2.column_dimensions[col_letter].width = max(
                    min(max_len + 4, 50), min_w)

            # ── Legend footer ─────────────────────────────
            tat_footer_row = tat_header_row + len(discharged) + 2
            ws2.merge_cells(
                f"A{tat_footer_row}:{TAT_LAST}{tat_footer_row}")
            tf                = ws2[f"A{tat_footer_row}"]
            tf.value          = (
                "TAT LEGEND:  "
                "🟢 GREEN = ≤ 30 minutes (Fast)  |  "
                "🟡 ORANGE = 31–60 minutes (Acceptable)  |  "
                "🔴 RED = > 60 minutes (Slow)  |  "
                "🟣 PURPLE = Negative TAT (check data)"
            )
            tf.font           = _font(italic=True, color=XL_MUTED, size=8)
            tf.fill           = _fill(XL_HEADING_BG)
            tf.alignment      = _align("center")
            tf.border         = _border()

            # ── Freeze & filter ───────────────────────────
            ws2.freeze_panes    = "A6"
            ws2.auto_filter.ref = f"A{tat_header_row}:{TAT_LAST}{tat_header_row}"

            # ── Sheet protection ──────────────────────────
            for sheet in [ws, ws2]:
                sheet.protection.sheet               = True
                sheet.protection.password            = "qphn2025"
                sheet.protection.selectLockedCells   = False
                sheet.protection.selectUnlockedCells = False
                sheet.protection.formatCells         = True
                sheet.protection.insertRows          = True
                sheet.protection.deleteRows          = True
                sheet.protection.sort                = True
                sheet.protection.autoFilter          = False

            # ── Print settings ────────────────────────────
            ws2.page_setup.orientation = "landscape"
            ws2.page_setup.fitToPage   = True
            ws2.page_setup.fitToWidth  = 1
            ws2.page_setup.fitToHeight = 0

            wb.save(file_path)
            messagebox.showinfo("Export Complete",
                f"ER report saved to:\n{file_path}\n\n"
                f"Sheet 1 — ER Patient Report:     {len(patients)} record(s)\n"
                f"Sheet 2 — Discharge TAT Report:  {len(discharged)} discharged")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report:\n{e}")

    # ── Buttons ───────────────────────────────────────────
    Label(content,
          text="Export all ER patient registration data to a styled Excel worksheet.",
          font=FONT["body"], bg=T["bg"], fg=T["text"]).pack(anchor=W, pady=(0, 4))

    status_lbl = Label(content, text="", font=FONT["small"],
                       bg=T["bg"], fg=T["muted"])
    status_lbl.pack(anchor=W, pady=(0, 12))

    def refresh_with_status():
        if refresh_database():
            status_lbl.config(
                text=f"✓ Data refreshed at {datetime.now().strftime('%H:%M:%S')}",
                fg=T["success"])
            status_lbl.after(3000, lambda: status_lbl.config(
                text="Database ready.", fg=T["muted"]))

    bf = Frame(content, bg=T["bg"])
    bf.pack(anchor=W)
    Button(bf, text="Export ER Report", command=export_to_excel,
           bg=T["accent"], fg=T["white"], font=FONT["body_b"],
           bd=0, relief=FLAT, padx=14, pady=10,
           cursor="hand2").pack(side=LEFT, padx=(0, 10))
    Button(bf, text="Refresh Data", command=refresh_with_status,
           bg=T["border2"], fg=T["text"], font=FONT["body_b"],
           bd=0, relief=FLAT, padx=14, pady=10,
           cursor="hand2").pack(side=LEFT)

    refresh_with_status()

    if page_refreshers is not None:
        page_refreshers["er_report"] = lambda: build_er_report_page(
            parent, page_refreshers)

    return frame